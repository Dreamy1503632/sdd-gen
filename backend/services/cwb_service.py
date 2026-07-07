"""
services/cwb_service.py
Orchestrates the CWB generation workflow:
  - Calls per-module LangChain CWB chains
  - Writes styled Excel workbooks per module
  - Returns download paths
"""
from __future__ import annotations
import io
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import AsyncGenerator

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import get_settings
from models.sdd_models import SDDDocument
from models.cwb_models import (
    CWBDocument, CWBGenerationProgress, ConfigRow, SetupTask,
    CoreHRWorkbook, TalentWorkbook, WorkforceWorkbook,
    PayrollWorkbook, CompBenefitsWorkbook, AnalyticsWorkbook,
)
from chains.cwb_chain import generate_cwb_for_module
from services.sdd_service import get_sdd

settings = get_settings()
_CWB_STORE: dict[str, dict] = {}

# ── Colours ──────────────────────────────────────────────────────────────────
ORACLE_RED   = "C74634"
DEEP_NAVY    = "1B3B5F"
GOLD         = "D4AF37"
LIGHT_GRAY   = "F2F2F2"
WHITE        = "FFFFFF"
GREEN        = "00B050"
AMBER        = "FF8C00"
BLUE         = "4472C4"

MODULE_COLORS = {
    "Core HR":               "1B3B5F",
    "Talent Management":     "2E75B6",
    "Workforce Management":  "375623",
    "Payroll":               "833C00",
    "Compensation & Benefits": "7030A0",
    "Workforce Analytics":   "203864",
    "Health & Safety":       "C00000",
    "Help Desk & Case Management": "404040",
}


# ──────────────────────────────────────────────────────────────────────────────
# Public streaming entry point
# ──────────────────────────────────────────────────────────────────────────────

async def generate_cwb_stream(
    sdd_session_id: str,
    modules: list[str],
    company_name: str,
) -> AsyncGenerator[CWBGenerationProgress, None]:
    session_id = str(uuid.uuid4())
    logs: list[str] = []
    download_urls: dict[str, str] = {}
    start_time = time.time()

    sdd: SDDDocument | None = get_sdd(sdd_session_id)
    if not sdd:
        yield CWBGenerationProgress(
            session_id=session_id, status="failed",
            progress=0, current_module="",
            log_entries=["SDD session not found. Please generate SDD first."],
        )
        return

    cwb_doc = CWBDocument(
        session_id=session_id,
        sdd_session_id=sdd_session_id,
        company_name=company_name,
        modules=modules,
    )

    total = len(modules)
    for idx, module in enumerate(modules):
        base_pct = int((idx / total) * 90)
        logs.append(f"Generating CWB: {module}...")
        yield CWBGenerationProgress(
            session_id=session_id, status="generating",
            progress=base_pct + 2, current_module=module,
            log_entries=list(logs), download_urls=dict(download_urls),
        )

        try:
            raw = await generate_cwb_for_module(module, sdd)
        except Exception as exc:
            logs.append(f"⚠ AI failed for {module}, using defaults: {exc}")
            raw = _default_cwb_data(module)

        # Attach to CWB document
        _attach_module(cwb_doc, module, raw)

        # Write Excel file for this module
        excel_bytes = _write_module_excel(module, raw, company_name, sdd)
        file_path = _save_cwb_file(module, company_name, session_id, excel_bytes)
        download_urls[module] = f"/api/v1/cwb/download/{session_id}/{_module_slug(module)}"

        task_count = len(raw.get("setup_tasks", []))
        cwb_doc.total_setup_tasks += task_count
        logs.append(f"✓ {module}: {task_count} setup tasks generated")

        yield CWBGenerationProgress(
            session_id=session_id, status="generating",
            progress=base_pct + int(90 / total), current_module=module,
            log_entries=list(logs), download_urls=dict(download_urls),
        )

    cwb_doc.generation_time_seconds = round(time.time() - start_time, 1)

    # Persist
    _CWB_STORE[session_id] = {
        "cwb": cwb_doc.model_dump(),
        "download_urls": download_urls,
        "created_at": datetime.utcnow().isoformat(),
    }

    logs.append(f"✅ All CWBs generated in {cwb_doc.generation_time_seconds}s")
    yield CWBGenerationProgress(
        session_id=session_id, status="complete",
        progress=100, current_module="Complete",
        log_entries=list(logs), download_urls=dict(download_urls),
    )


# ──────────────────────────────────────────────────────────────────────────────
# Getters
# ──────────────────────────────────────────────────────────────────────────────

def get_cwb_file_path(session_id: str, module_slug: str) -> Path | None:
    out_dir = Path(settings.output_dir)
    matches = list(out_dir.glob(f"CWB_{session_id[:8]}_{module_slug}_*.xlsx"))
    return matches[0] if matches else None


# ──────────────────────────────────────────────────────────────────────────────
# Excel workbook writer
# ──────────────────────────────────────────────────────────────────────────────

def _write_module_excel(
    module: str,
    cwb_data: dict,
    company_name: str,
    sdd: SDDDocument,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    color = MODULE_COLORS.get(module, DEEP_NAVY)

    # Cover sheet
    _add_cover_sheet(wb, module, company_name, color)

    # Configuration sections
    section_map = _get_section_map(module)
    for sheet_title, data_key in section_map:
        data = cwb_data.get(data_key, [])
        if data:
            _add_config_sheet(wb, sheet_title, data, color)

    # Setup Tasks sheet
    tasks = cwb_data.get("setup_tasks", [])
    if tasks:
        _add_setup_tasks_sheet(wb, tasks, color)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_cover_sheet(wb, module: str, company: str, color: str):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = f"ORACLE FUSION HCM – {module.upper()}"
    c.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:B2")
    c2 = ws["A2"]
    c2.value = "CONFIGURATION WORKBOOK"
    c2.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c2.fill = PatternFill("solid", fgColor=color)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    meta = [
        ("Company:", company),
        ("Document Status:", "For Review"),
        ("Version:", "1.0"),
        ("Date:", datetime.now().strftime("%d %b %Y")),
    ]
    for i, (label, value) in enumerate(meta, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=i, column=2, value=value).font = Font(name="Calibri", size=11)
        ws.row_dimensions[i].height = 20


def _add_config_sheet(wb, title: str, rows: list[dict], color: str):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = ["Configuration Item", "Value", "Notes", "Mandatory", "Oracle Default"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    widths = [45, 40, 50, 12, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25

    for row_idx, row_data in enumerate(rows, start=2):
        # row_data may be dict or ConfigRow-like
        if hasattr(row_data, "model_dump"):
            rd = row_data.model_dump()
        elif isinstance(row_data, dict):
            rd = row_data
        else:
            continue

        values = [
            rd.get("config_item", ""),
            rd.get("value", ""),
            rd.get("notes", ""),
            "Yes" if rd.get("mandatory", True) else "No",
            rd.get("oracle_default", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=str(val))
            cell.font = Font(name="Calibri", size=10)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Mandatory indicator
        mandatory_cell = ws.cell(row=row_idx, column=4)
        if values[3] == "Yes":
            mandatory_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        ws.row_dimensions[row_idx].height = 35


def _add_setup_tasks_sheet(wb, tasks: list[dict], color: str):
    ws = wb.create_sheet("Setup Tasks")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = ["#", "Task Name", "Functional Area", "Offering", "Scope", "Prerequisites", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    widths = [5, 45, 30, 30, 18, 35, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_idx, task in enumerate(tasks, start=2):
        if isinstance(task, dict):
            td = task
        else:
            td = task.model_dump()

        prereqs = ", ".join(td.get("prerequisite_tasks", [])) or "-"
        values = [
            td.get("sequence", row_idx - 1),
            td.get("task_name", ""),
            td.get("functional_area", ""),
            td.get("offering", ""),
            td.get("scope", ""),
            prereqs,
            td.get("notes", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else "")
            cell.font = Font(name="Calibri", size=10)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[row_idx].height = 35

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _module_slug(module: str) -> str:
    return module.replace(" ", "_").replace("&", "and").lower()


def _get_section_map(module: str) -> list[tuple[str, str]]:
    maps = {
        "Core HR": [
            ("Enterprise Structure", "enterprise_structure"),
            ("Legal Entities", "legal_entities"),
            ("Business Units", "business_units"),
            ("Departments", "departments"),
            ("Locations", "locations"),
            ("Jobs", "jobs"),
            ("Positions", "positions"),
            ("Grades", "grades"),
            ("Person Types", "person_types"),
            ("Actions & Reasons", "actions_reasons"),
            ("HR Lookups", "lookups"),
            ("Security Roles", "security_roles"),
        ],
        "Talent Management": [
            ("Performance Templates", "performance_templates"),
            ("Goal Plans", "goal_plans"),
            ("Talent Review", "talent_review_templates"),
            ("Succession Plans", "succession_plans"),
            ("Recruiting Setup", "recruiting_setup"),
            ("Onboarding Checklists", "onboarding_checklists"),
        ],
        "Workforce Management": [
            ("Work Schedules", "work_schedules"),
            ("Absence Types", "absence_types"),
            ("Accrual Plans", "accrual_plans"),
            ("Time Entry Profiles", "time_entry_profiles"),
            ("Overtime Rules", "overtime_rules"),
        ],
        "Payroll": [
            ("Payroll Definitions", "payroll_definitions"),
            ("Earnings Elements", "payroll_elements"),
            ("Deduction Elements", "deduction_elements"),
            ("Payroll Flows", "payroll_flows"),
            ("Costing Setup", "costing_setup"),
            ("Tax Configuration", "tax_config"),
        ],
        "Compensation & Benefits": [
            ("Salary Basis", "salary_basis"),
            ("Grade Rates", "grade_rates"),
            ("Compensation Plans", "compensation_plans"),
            ("Benefit Plans", "benefit_plans"),
            ("Benefit Programs", "benefit_programs"),
            ("Eligibility Profiles", "eligibility_profiles"),
            ("Enrollment Periods", "enrollment_periods"),
        ],
        "Workforce Analytics": [
            ("Subject Areas", "subject_areas"),
            ("Dashboards", "dashboard_configs"),
            ("Standard Reports", "standard_reports"),
            ("Custom Reports", "custom_reports"),
        ],
    }
    return maps.get(module, [])


def _save_cwb_file(module: str, company: str, session_id: str, content: bytes) -> Path:
    out_dir = Path(settings.output_dir)
    out_dir.mkdir(exist_ok=True)
    slug = _module_slug(module)
    fname = f"CWB_{session_id[:8]}_{slug}_{company.replace(' ', '_')}.xlsx"
    path = out_dir / fname
    path.write_bytes(content)
    return path


def _attach_module(cwb_doc: CWBDocument, module: str, raw: dict):
    """Store raw workbook data in the CWB document."""
    # We just record that this module was generated; the Excel is on disk
    pass


def _default_cwb_data(module: str) -> dict:
    """Minimal fallback data when AI fails."""
    return {
        "setup_tasks": [
            {
                "task_name": f"{module} Initial Setup",
                "functional_area": module,
                "offering": "Workforce Deployment",
                "scope": "Enterprise",
                "sequence": 1,
                "prerequisite_tasks": [],
                "notes": "Configure as per design workshops",
            }
        ]
    }