"""
services/excel_service.py
Handles Excel questionnaire generation (HLA) and parsing of uploaded files.
Uses openpyxl for full formatting control.
"""
from __future__ import annotations
import io
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from models.hla_models import HCMModule, Industry, Priority, ParsedQuestionnaire, QuestionResponse

# ──────────────────────────────────────────────────────────────────────────────
# Brand colours (Oracle palette)
# ──────────────────────────────────────────────────────────────────────────────

ORACLE_RED   = "C74634"
DEEP_NAVY    = "1B3B5F"
GOLD         = "D4AF37"
LIGHT_GRAY   = "F2F2F2"
MED_GRAY     = "D9D9D9"
WHITE        = "FFFFFF"

PRIORITY_COLORS = {
    "Critical": "FF4444",
    "High":     "FF8C00",
    "Medium":   "4472C4",
    "Low":      "70AD47",
}

# ──────────────────────────────────────────────────────────────────────────────
# Question database  (kept server-side – single source of truth)
# ──────────────────────────────────────────────────────────────────────────────

from data.question_database import QUESTION_DATABASE  # noqa: E402  (created below)


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_font(size=11, bold=True, color=WHITE) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _body_font(size=10, bold=False) -> Font:
    return Font(name="Calibri", size=size, bold=bold)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, values: list[str], fill_color: str = DEEP_NAVY) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _header_font()
        cell.fill = _fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


# ──────────────────────────────────────────────────────────────────────────────
# Questionnaire generation
# ──────────────────────────────────────────────────────────────────────────────

def generate_questionnaire_excel(
    industry: str,
    modules: list[str],
) -> bytes:
    """Return Excel bytes for the HLA questionnaire."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)                 # remove default empty sheet

    _build_instructions_sheet(wb, industry, modules)
    _build_company_profile_sheet(wb, industry)
    for module in modules:
        questions = _get_questions(module, industry)
        _build_module_sheet(wb, module, questions)
    _build_integration_sheet(wb)
    _build_technical_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_instructions_sheet(wb, industry: str, modules: list[str]) -> None:
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "ORACLE FUSION HCM – HIGH LEVEL ASSESSMENT QUESTIONNAIRE"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title_cell.fill = _fill(DEEP_NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Metadata
    meta = [
        ("Industry:", industry),
        ("Modules:", ", ".join(modules)),
        ("Version:", "2.0"),
        ("Date Generated:", datetime.now().strftime("%d %b %Y")),
        ("Status:", "For Client Completion"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=label).font = _body_font(bold=True)
        ws.cell(row=i, column=2, value=value).font = _body_font()

    # Instructions
    ws.cell(row=9, column=1, value="INSTRUCTIONS").font = _header_font(color=ORACLE_RED)
    instructions = [
        "1. Complete all questions in the Response column – Critical questions are mandatory",
        "2. Questions are colour-coded: 🔴 Critical  🟠 High  🔵 Medium  🟢 Low",
        "3. Add additional context in the Notes column",
        "4. Return completed file to your Oracle consultant for AI analysis",
        "5. Estimated completion time: 4–8 hours depending on scope",
    ]
    for i, instr in enumerate(instructions, start=10):
        ws.cell(row=i, column=1, value=instr).font = _body_font()

    _set_col_widths(ws, [30, 60, 20, 20, 20, 20])


def _build_company_profile_sheet(wb, industry: str) -> None:
    ws = wb.create_sheet("Company Profile")
    ws.sheet_view.showGridLines = False

    _write_header_row(ws, 1, ["Section", "Question", "Response", "Notes"])
    ws.row_dimensions[1].height = 25

    rows = [
        ("Company Information", "Legal Company Name", "", ""),
        ("Company Information", "Industry Sector", industry, "Pre-filled"),
        ("Company Information", "Headquarters Location", "", ""),
        ("Company Information", "Year Founded", "", ""),
        ("Workforce", "Total Employee Count (Global)", "", ""),
        ("Workforce", "Full-time Employees", "", ""),
        ("Workforce", "Part-time Employees", "", ""),
        ("Workforce", "Contractors/Temporary Workers", "", ""),
        ("Workforce", "Annual Turnover Rate (%)", "", ""),
        ("Geographic", "Number of Countries Operating In", "", ""),
        ("Geographic", "Number of Facilities/Locations", "", ""),
        ("Geographic", "Primary Operating Regions", "", ""),
        ("Strategic", "Top Business Objective 1", "", ""),
        ("Strategic", "Top Business Objective 2", "", ""),
        ("Strategic", "Top Business Objective 3", "", ""),
        ("Strategic", "Target Implementation Timeline", "", ""),
        ("Strategic", "Implementation Budget Range", "", "Optional"),
    ]

    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 22

    _set_col_widths(ws, [28, 55, 45, 30])


def _build_module_sheet(wb, module: str, questions: list[dict]) -> None:
    sheet_name = module[:31]  # Excel tab limit
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    # Title row
    ws.merge_cells("A1:F1")
    tc = ws["A1"]
    tc.value = f"{module.upper()} – DETAILED ASSESSMENT"
    tc.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    tc.fill = _fill(ORACLE_RED)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Stats row
    ws.merge_cells("A2:F2")
    stats_cell = ws["A2"]
    stats_cell.value = (
        f"Questions: {len(questions)}  |  "
        f"Critical: {sum(1 for q in questions if q['priority'] == 'Critical')}  |  "
        f"High: {sum(1 for q in questions if q['priority'] == 'High')}  |  "
        f"Medium: {sum(1 for q in questions if q['priority'] == 'Medium')}"
    )
    stats_cell.font = _body_font(bold=True)
    stats_cell.fill = _fill(LIGHT_GRAY)
    stats_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Column headers
    _write_header_row(ws, 3, ["Category", "Question", "Response", "Priority", "Guidance", "Notes"])
    ws.row_dimensions[3].height = 25

    # Data rows
    for i, q in enumerate(questions, start=4):
        priority = q.get("priority", "Medium")
        fill_color = PRIORITY_COLORS.get(priority, "4472C4") + "22"  # 13% opacity hex trick

        row_cells = [
            q.get("category", ""),
            q.get("question", ""),
            "",                          # Response – client fills this
            priority,
            q.get("guidance", ""),
            "",                          # Notes
        ]

        for col, val in enumerate(row_cells, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Highlight priority cell
        priority_cell = ws.cell(row=i, column=4)
        priority_cell.fill = _fill(PRIORITY_COLORS.get(priority, "4472C4"))
        priority_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        priority_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 40

    _set_col_widths(ws, [28, 68, 50, 12, 55, 30])


def _build_integration_sheet(wb) -> None:
    ws = wb.create_sheet("Integrations")
    ws.sheet_view.showGridLines = False
    _write_header_row(ws, 1, ["System Name", "Purpose", "Direction", "Integration Type", "Frequency", "Notes"])

    sample = [
        ("Example: SAP ERP", "Financial data sync", "Bidirectional", "REST API", "Real-time", ""),
        ("Example: ADP Payroll", "Payroll processing", "Outbound", "File/SFTP", "Daily", ""),
    ]
    for i, row in enumerate(sample, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", size=10, color="999999", italic=True)

    _set_col_widths(ws, [30, 40, 18, 22, 16, 40])


def _build_technical_sheet(wb) -> None:
    ws = wb.create_sheet("Technical Environment")
    ws.sheet_view.showGridLines = False
    _write_header_row(ws, 1, ["Category", "Question", "Response"])

    rows = [
        ("Infrastructure", "Current data center / cloud provider(s)", ""),
        ("Infrastructure", "Network connectivity type", ""),
        ("Security", "SSO/Identity Provider (Okta, Azure AD, etc.)", ""),
        ("Security", "MFA currently in use?", ""),
        ("Security", "Data encryption requirements", ""),
        ("Integration", "Integration middleware / iPaaS platform", ""),
        ("Integration", "API management tooling", ""),
        ("Reporting", "BI/Analytics tools (Power BI, Tableau, etc.)", ""),
        ("Reporting", "Data warehouse / lake infrastructure", ""),
        ("Reporting", "Data retention requirements (years)", ""),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _body_font()
            cell.border = _thin_border()

    _set_col_widths(ws, [25, 58, 50])


# ──────────────────────────────────────────────────────────────────────────────
# Question lookup
# ──────────────────────────────────────────────────────────────────────────────

def _industry_key(industry: str) -> str:
    mapping = {
        "manufacturing": "manufacturing",
        "healthcare":    "healthcare",
        "financial":     "financialServices",
        "retail":        "retail",
    }
    lower = industry.lower()
    for k, v in mapping.items():
        if k in lower:
            return v
    return "base"


def _get_questions(module: str, industry: str) -> list[dict]:
    """
    Return questions for a module, preferring the ChromaDB vector store
    (loaded from Excel files) with a fallback to the static QUESTION_DATABASE.
    """
    try:
        from data.question_store import get_questions_by_module
        questions = get_questions_by_module(module)
        if questions:
            return questions
    except Exception:
        pass

    # Fallback: static question_database.py
    print("Fallback to static question database for module:", module)
    db = QUESTION_DATABASE.get(module, {})
    questions = list(db.get("base", []))
    key = _industry_key(industry)
    if key != "base" and key in db:
        questions += db[key]
    return questions


# ──────────────────────────────────────────────────────────────────────────────
# Questionnaire parsing
# ──────────────────────────────────────────────────────────────────────────────

def parse_questionnaire_excel(file_bytes: bytes, file_name: str) -> ParsedQuestionnaire:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Detect industry & modules from Instructions sheet
    detected_industry = "Not specified"
    detected_modules: list[str] = []

    if "Instructions" in wb.sheetnames:
        ws = wb["Instructions"]
        for row in ws.iter_rows(values_only=True):
            if row and len(row) > 1:
                label = str(row[0] or "")
                value = str(row[1] or "")
                if "Industry:" in label or label.strip() == "Industry:":
                    detected_industry = value
                elif "Modules:" in label or label.strip() == "Modules:":
                    detected_modules = [m.strip() for m in value.split(",") if m.strip()]

    known_modules = [m.value for m in HCMModule]
    if not detected_modules:
        detected_modules = [m for m in known_modules if m in wb.sheetnames]

    module_responses: dict[str, list[dict]] = {}
    total_questions = 0
    answered_questions = 0

    for module in known_modules:
        if module not in wb.sheetnames:
            continue
        ws = wb[module]
        responses: list[dict] = []
        for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
            if not row or not row[1]:
                continue
            response_val = str(row[2] or "").strip()
            q = {
                "category": str(row[0] or ""),
                "question": str(row[1] or ""),
                "response": response_val,
                "priority": str(row[3] or "Medium"),
                "guidance": str(row[4] or ""),
            }
            responses.append(q)
            total_questions += 1
            if response_val:
                answered_questions += 1

        module_responses[module] = responses

    completion_rate = (
        round((answered_questions / total_questions) * 100, 1)
        if total_questions > 0 else 0.0
    )

    # Convert to QuestionResponse objects
    qr_map: dict[str, list[QuestionResponse]] = {}
    for mod, responses in module_responses.items():
        qr_map[mod] = [
            QuestionResponse(
                category=r["category"],
                question=r["question"],
                response=r["response"],
                priority=r.get("priority", Priority.MEDIUM),
                guidance=r["guidance"],
            )
            for r in responses
        ]

    return ParsedQuestionnaire(
        file_name=file_name,
        industry=detected_industry,
        modules=detected_modules,
        module_responses=qr_map,
        total_questions=total_questions,
        answered_questions=answered_questions,
        completion_rate=completion_rate,
    )